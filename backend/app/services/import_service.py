import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from app.models import TravelType, Client
from app.schemas import ImportError

HEADER_MAPS = {
    "en": {
        "surname": ["surname", "last name", "family name"],
        "given_name": ["given name", "first name", "name"],
        "father_name": ["father name", "father's name"],
        "mother_name": ["mother name", "mother's name"],
        "passport_number": ["passport number", "passport no", "passport"],
        "nationality": ["nationality"],
        "date_of_birth": ["date of birth", "dob", "birth date"],
        "passport_issue_date": ["passport issue date", "issue date"],
        "passport_expiry": ["passport expiry", "expiry date", "passport expiration"],
        "gender": ["gender", "sex"],
        "travel_type": ["travel type", "type of travel"],
        "payment_method": ["payment method", "payment"],
        "travel_date": ["travel date", "departure date"],
        "notes": ["notes", "remarks", "comments"],
    },
    "fr": {
        "surname": ["nom", "nom de famille"],
        "given_name": ["prénom", "prénom(s)"],
        "father_name": ["nom du père", "père"],
        "mother_name": ["nom de la mère", "mère"],
        "passport_number": ["n° passeport", "numéro de passeport", "passeport"],
        "nationality": ["nationalité"],
        "date_of_birth": ["date de naissance", "naissance"],
        "passport_issue_date": ["date d'émission", "date de délivrance"],
        "passport_expiry": ["date d'expiration", "expiration"],
        "gender": ["genre", "sexe"],
        "travel_type": ["type de voyage", "voyage"],
        "payment_method": ["mode de paiement", "paiement"],
        "travel_date": ["date de voyage", "départ"],
        "notes": ["remarques", "notes", "commentaires"],
    },
    "ar": {
        "surname": ["اللقب", "اسم العائلة"],
        "given_name": ["الاسم", "الاسم الأول"],
        "father_name": ["اسم الأب", "الأب"],
        "mother_name": ["اسم الأم", "الأم"],
        "passport_number": ["رقم جواز السفر", "جواز السفر"],
        "nationality": ["الجنسية"],
        "date_of_birth": ["تاريخ الميلاد", "الميلاد"],
        "passport_issue_date": ["تاريخ الإصدار", "الإصدار"],
        "passport_expiry": ["تاريخ الانتهاء", "الانتهاء", "تاريخ الصلاحية"],
        "gender": ["الجنس", "النوع"],
        "travel_type": ["نوع السفر", "السفر"],
        "payment_method": ["طريقة الدفع", "الدفع"],
        "travel_date": ["تاريخ السفر", "السفر"],
        "notes": ["ملاحظات", "تعليقات"],
    }
}

def detect_language(headers: List[str]) -> str:
    scores = {"en": 0, "fr": 0, "ar": 0}
    h_lower = [h.strip().lower() for h in headers]
    for lang, mapping in HEADER_MAPS.items():
        for field, keywords in mapping.items():
            for kw in keywords:
                if kw in h_lower:
                    scores[lang] += 1
    return max(scores, key=scores.get) if max(scores.values()) > 0 else "en"

def map_headers(headers: List[str], lang: str) -> Dict[str, int]:
    h_lower = [h.strip().lower() for h in headers]
    mapping = {}
    for field, keywords in HEADER_MAPS.get(lang, HEADER_MAPS["en"]).items():
        for idx, h in enumerate(h_lower):
            if any(kw in h for kw in keywords):
                mapping[field] = idx
                break
    return mapping

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None

def normalize_gender(val, lang="en"):
    if not val:
        return None
    v = str(val).strip().upper()
    if v in ("M", "MALE", "HOMME", "ذكر"):
        return "M"
    if v in ("F", "FEMALE", "FEMME", "أنثى"):
        return "F"
    return None

def parse_import_file(content: bytes, filename: str) -> Tuple[List[Dict[str, Any]], str]:
    rows = []
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        headers = next(reader)
        lang = detect_language(headers)
        col_map = map_headers(headers, lang)
        for r in reader:
            if not any(r):
                continue
            row = {}
            for field, idx in col_map.items():
                if idx < len(r):
                    row[field] = r[idx].strip()
            rows.append(row)
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value) if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        lang = detect_language(headers)
        col_map = map_headers(headers, lang)
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            row = {}
            for field, idx in col_map.items():
                if idx < len(r):
                    row[field] = r[idx]
            rows.append(row)
    return rows, lang

async def validate_rows(rows: List[Dict], db: AsyncSession) -> Tuple[List[Dict], List[ImportError]]:
    errors = []
    valid = []
    # Fetch travel types for validation
    result = await db.execute(select(TravelType).where(TravelType.is_active == True))
    travel_types = {t.code: t.id for t in result.scalars().all()}
    tt_names = {}
    for t in result.scalars().all():
        for name in (t.name_en, t.name_fr, t.name_ar):
            tt_names[name.lower()] = t.id

    # Fetch existing passports
    passports = [str(r.get("passport_number", "")).strip() for r in rows if r.get("passport_number")]
    existing = set()
    if passports:
        res = await db.execute(select(Client.passport_number).where(
            and_(Client.passport_number.in_(passports), Client.archived == False)
        ))
        existing = {p for p in res.scalars().all()}

    for idx, row in enumerate(rows, start=2):
        row_errors = []
        # Required fields
        for req in ["surname", "given_name", "father_name", "passport_number", "nationality", "travel_type", "travel_date"]:
            if not row.get(req):
                row_errors.append(ImportError(row=idx, field=req, error="required"))

        # Passport uniqueness
        pp = str(row.get("passport_number", "")).strip()
        if pp and pp in existing:
            row_errors.append(ImportError(row=idx, field="passport_number", error="duplicate"))

        # Gender
        g = normalize_gender(row.get("gender"))
        if row.get("gender") and not g:
            row_errors.append(ImportError(row=idx, field="gender", error="invalid"))
        else:
            row["gender"] = g

        # Dates
        for dfield in ["date_of_birth", "passport_issue_date", "passport_expiry", "travel_date"]:
            if row.get(dfield):
                parsed = parse_date(row[dfield])
                if not parsed:
                    row_errors.append(ImportError(row=idx, field=dfield, error="invalid_date"))
                else:
                    row[dfield] = parsed

        # Travel type
        tt_val = str(row.get("travel_type", "")).strip().lower()
        tt_id = travel_types.get(tt_val) or tt_names.get(tt_val)
        if not tt_id and row.get("travel_type"):
            row_errors.append(ImportError(row=idx, field="travel_type", error="invalid_code"))
        else:
            row["travel_type_id"] = tt_id

        if not row_errors:
            valid.append(row)
        else:
            errors.extend(row_errors)

    return valid, errors
