import csv
import io
import json
import uuid
from datetime import date, datetime
from typing import List

from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models import Client
from app.dependencies import redis_client

REQUIRED_COLUMNS = {
    "surname", "given_name", "father_name", "passport_number",
    "nationality", "travel_type_id", "travel_date",
}


def _parse_date(val: str) -> date | None:
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return None


async def parse_and_validate(file_content: bytes, filename: str) -> dict:
    validation_id = str(uuid.uuid4())
    errors = []
    rows = []

    try:
        if filename.endswith(".csv"):
            text = file_content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            raw_rows = list(reader)
        elif filename.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            col_headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append(dict(zip(col_headers, [str(v or "").strip() for v in row])))
        else:
            return {"validation_id": validation_id, "total_rows": 0, "valid_rows": 0, "errors": [{"row": 0, "message": "Unsupported file format"}]}

        for i, row in enumerate(raw_rows, start=2):
            row_errors = []
            # Normalize keys
            row = {k.lower().strip(): v for k, v in row.items()}
            for col in REQUIRED_COLUMNS:
                if col not in row or not row[col]:
                    row_errors.append(f"Missing required field: {col}")

            travel_date = _parse_date(row.get("travel_date", ""))
            if row.get("travel_date") and not travel_date:
                row_errors.append("Invalid travel_date format (use YYYY-MM-DD)")

            gender = row.get("gender", "").upper()
            if gender and gender not in ("M", "F"):
                row_errors.append("Gender must be M or F")

            if row_errors:
                errors.append({"row": i, "messages": row_errors})
            else:
                rows.append({
                    "surname": row["surname"],
                    "given_name": row["given_name"],
                    "father_name": row["father_name"],
                    "mother_name": row.get("mother_name") or None,
                    "passport_number": row["passport_number"],
                    "nationality": row["nationality"],
                    "date_of_birth": str(_parse_date(row.get("date_of_birth", ""))) if row.get("date_of_birth") else None,
                    "passport_issue_date": str(_parse_date(row.get("passport_issue_date", ""))) if row.get("passport_issue_date") else None,
                    "passport_expiry": str(_parse_date(row.get("passport_expiry", ""))) if row.get("passport_expiry") else None,
                    "gender": gender or None,
                    "travel_type_id": int(row["travel_type_id"]),
                    "payment_method": row.get("payment_method", "cash"),
                    "travel_date": str(travel_date),
                    "notes": row.get("notes") or None,
                })

    except Exception as exc:
        errors.append({"row": 0, "messages": [f"Parse error: {exc}"]})

    # Cache validated rows in Redis for 10 minutes
    await redis_client.setex(
        f"import:{validation_id}",
        600,
        json.dumps({"rows": rows}),
    )

    return {
        "validation_id": validation_id,
        "total_rows": len(raw_rows) if "raw_rows" in dir() else 0,
        "valid_rows": len(rows),
        "errors": errors,
    }


async def commit_import(validation_id: str, rows: list) -> dict:
    async with AsyncSessionLocal() as session:
        passports = [r["passport_number"] for r in rows]
        existing = set(
            (await session.execute(
                select(Client.passport_number).where(Client.passport_number.in_(passports))
            )).scalars().all()
        )
        new_clients = []
        skipped = 0
        for row in rows:
            if row["passport_number"] in existing:
                skipped += 1
                continue
            client = Client(
                surname=row["surname"],
                given_name=row["given_name"],
                father_name=row["father_name"],
                mother_name=row.get("mother_name"),
                passport_number=row["passport_number"],
                nationality=row["nationality"],
                gender=row.get("gender"),
                travel_type_id=int(row["travel_type_id"]),
                payment_method=row.get("payment_method", "cash"),
                travel_date=_parse_date(row["travel_date"]),
                date_of_birth=_parse_date(row.get("date_of_birth", "")),
                passport_issue_date=_parse_date(row.get("passport_issue_date", "")),
                passport_expiry=_parse_date(row.get("passport_expiry", "")),
                notes=row.get("notes"),
            )
            session.add(client)
            new_clients.append(client)
        await session.commit()
        # Invalidate cache
        await redis_client.delete(f"import:{validation_id}")
        return {"imported": len(new_clients), "skipped": skipped}
